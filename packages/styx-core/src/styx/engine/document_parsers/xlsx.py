"""XLSX parser через openpyxl (волна 28).

Pure-Python чтение Excel XML. Каждая вкладка → блок с заголовком
``=== Sheet: <name> ===`` + строки cells через `\t`. Формулы не
вычисляются (`data_only=True` берёт cached values).
"""

from __future__ import annotations

from pathlib import Path

from styx.engine.document_parsers._types import ParsedDocument


def parse_xlsx(path: Path) -> ParsedDocument:
    """Извлечь текст из .xlsx.

    Raises:
        ValueError: malformed XLSX (openpyxl не распарсил).
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except (InvalidFileException, KeyError) as exc:
        raise ValueError(f"XLSX parse error: {exc}") from exc

    parts: list[str] = []
    sheet_names: list[str] = []
    total_cells = 0
    for ws in wb.worksheets:
        sheet_names.append(ws.title)
        parts.append(f"=== Sheet: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells).rstrip())
                total_cells += sum(1 for c in cells if c.strip())

    wb.close()

    text = "\n".join(parts).strip()

    return ParsedDocument(
        text=text,
        metadata={
            "sheet_names": sheet_names,
            "non_empty_cells": total_cells,
        },
    )
